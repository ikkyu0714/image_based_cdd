"""
https://aidiary.hatenablog.com/entry/20170110/1484057655 参考
"""
#from re import T
from tensorflow.keras.applications import VGG16
from tensorflow.keras.applications.vgg16 import VGG16, decode_predictions, preprocess_input
import os, urllib.request, collections
import numpy as np
import pandas as pd
from sklearn.cluster import KMeans
from sklearn.cluster import DBSCAN
from sklearn.ensemble import IsolationForest
from scipy.spatial import distance
from keras.preprocessing.image import ImageDataGenerator, load_img, img_to_array
from PIL import Image
from tensorflow.python.eager.tape import delete_trace
from google_images_download_patch.google_images_download.google_images_download import googleimagesdownload
import openpyxl
import itertools
from PIL import ImageFile
import collections
from collections import defaultdict
ImageFile.LOAD_TRUNCATED_IMAGES = True

# 文化差検出クラス
class Cultural_Difference_Detector():
    def __init__(self):
        #self.readfile = 'synset_indonesiaのコピー2.xlsx' # 読み込むsynsetのリストを記録したファイル
        #self.readfile = 'synset_テスト.xlsx' # 読み込むsynsetのリストを記録したファイル
        self.readfile = 'synset_改良.xlsx' # 読み込むsynsetのリストを記録したファイルs
        #self.readfile = 'synsetアンケートのコピー.xlsx' # 読み込むsynsetのリストを記録したファイル
        #self.filename='vgg_predict_result.xlsx' # 結果を書き込むExcelのパス
        self.filename='../../研究/dbscan.xlsx' # 結果を書き込むExcelのパス
        self.jpn_keyword_dict = {} # キーワードのリストを取得する辞書（日本語）
        self.eng_keyword_dict = {} # キーワードのリストを取得する辞書（英語）
        self.ind_keyword_dict = {} #
        self.new_dir_path = "downloads" # グレースケール化した後のフォルダ
        self.dir_path = "downloads" # 画像が保存されているフォルダ

    # Synsetを取得する
    def get_synset(self):
        key = 1
        book = openpyxl.load_workbook(self.readfile)
        sheet = book.worksheets[0]

        for row in sheet.iter_rows(min_row=2, min_col=2):
            value = []
            for cell in row:
                #print(cell.value)
                values = cell.value
                values = values.replace(',', '')
                value.append(values.replace('_', ' '))

            self.jpn_keyword_dict[key] = value[1]
            self.eng_keyword_dict[key] = value[3]
            #self.ind_keyword_dict[key] = value[2]
            key += 1

        # 上書き保存（読み込んだのと同じ名前を指定）
        book.save(self.readfile)

        book.close()

    # 平均する関数(引数：リスト)
    def average(self, lists):
        total = sum(lists)
        average = total/(len(lists))
        return average

    # 結果をエクセルに書き込む
    def write_result(self, result, jpn_keyword, eng_keyword):
        book = openpyxl.load_workbook(self.filename)
        sheet = book.worksheets[0]
        max_row = sheet.max_row # シートの最後の行を取得

        # 最後の一つ下の行に (3列目:日本語のキーワード, 4列目:英語のキーワード, 5列目:cos類似度) を書き込む
        sheet.cell(row = max_row + 1, column=3).value = jpn_keyword
        sheet.cell(row = max_row + 1, column=4).value = eng_keyword
        sheet.cell(row = max_row + 1, column=5).value = result
        # 保存
        book.save(self.filename)
        # 終了
        book.close()

    # cos類似度を計算する
    def cos_sim(self, v1, v2):
        return np.dot(v1, v2) / (np.linalg.norm(v1) * np.linalg.norm(v2))

    # 画像をグレースケール化
    def imread_gray(self, image_name, path1, path2):
        gamma22LUT  = [pow(x/255.0, 2.2)*255 for x in range(256)] * 3
        gamma045LUT = [pow(x/255.0, 1.0/2.2)*255 for x in range(256)]
        img = Image.open(path1 + image_name)
        img_rgb = img.convert("RGB")
        img_rgbL = img_rgb.point(gamma22LUT)
        img_grayL = img_rgbL.convert("L")
        img_gray = img_grayL.point(gamma045LUT)
        img_gray.save(path2 + image_name)

    # 画像検索
    def image_search(self, keyword, exist_num):
        limit = 10 - exist_num
        response = googleimagesdownload()  #responseオブジェクトの生成
        arguments = {"keywords":keyword,  # 検索キーワード
                     "limit":limit,  # ダウンロードする画像の数(デフォルト100)
                     "format":"jpg" or "png"
                    }
        response.download(arguments)   # argumentsをresponseオブジェクトのdownloadメソッドに渡す

    # svgファイルを削除
    def svg_delete(self, jpn_keyword, eng_keyword):
        count = 0
        for filejpn in os.listdir(os.path.join(self.dir_path, jpn_keyword)):
            if filejpn.endswith('.svg') or filejpn.endswith('.webp'):
                count += 1
                print(os.path.join(os.path.join(self.dir_path, jpn_keyword), filejpn))
                os.remove(os.path.join(os.path.join(self.dir_path, jpn_keyword), filejpn))
        if count != 0:
            self.image_search(jpn_keyword, 10-count)
        count = 0
        for fileeng in os.listdir(os.path.join(self.dir_path, eng_keyword)):
            if fileeng.endswith('.svg') or filejpn.endswith('.webp'):
                count += 1
                print(os.path.join(os.path.join(self.dir_path, eng_keyword), fileeng))
                os.remove(os.path.join(os.path.join(self.dir_path, eng_keyword), fileeng))
        if count != 0:
            self.image_search(eng_keyword, 10-count)

    def print_cossim(self, cos_similarity):
        print("cos類似度は:",cos_similarity)

    # 結果を出力
    def print_result(self, c, feature_vector_average_dict, jpn_keyword, eng_keyword):
        num = 1
        print(jpn_keyword, eng_keyword)
        for i in feature_vector_average_dict:
            # 画像10枚分の特徴ベクトルの総和
            print(str(num)+".画像特徴量の総和\n",feature_vector_average_dict[i])
            # 平均特徴ベクトル
            feature_vector_average_dict[i] /= c[i]
            print(str(num)+".特徴量の平均値\n",feature_vector_average_dict[i])
            num += 1

        # cos類似度
        cos_similarity = self.cos_sim(feature_vector_average_dict["[1. 0.]"], feature_vector_average_dict["[0. 1.]"])
        print("cos類似度は:",cos_similarity)
        return cos_similarity

    # 画像があるか確認
    def image_confirm(self, jpn_keyword, eng_keyword):
        if os.path.isdir(os.path.join(self.dir_path, jpn_keyword)) == False:
            self.image_search(jpn_keyword, 0)
        else:
            files = os.listdir(os.path.join(self.dir_path, jpn_keyword))
            files_file = [f for f in files if os.path.isfile(os.path.join(os.path.join(self.dir_path, jpn_keyword), f))]
            if len(files_file) < 10:
                self.image_search(jpn_keyword, len(files_file))
        if os.path.isdir(os.path.join(self.dir_path, eng_keyword)) == False:
            self.image_search(eng_keyword, 0)
        else:
            files = os.listdir(os.path.join(self.dir_path, eng_keyword))
            files_file = [f for f in files if os.path.isfile(os.path.join(os.path.join(self.dir_path, eng_keyword), f))]
            if len(files_file) < 10:
                self.image_search(eng_keyword, len(files_file))
        """if os.path.isdir(os.path.join(self.dir_path, ind_keyword)) == False:
            self.image_search(ind_keyword, 0)
        else:
            files = os.listdir(os.path.join(self.dir_path, ind_keyword))
            files_file = [f for f in files if os.path.isfile(os.path.join(os.path.join(self.dir_path, ind_keyword), f))]
            if len(files_file) < 10:
                self.image_search(ind_keyword, len(files_file))"""

    # フォルダーがあるか確認
    def folder_confirm(self, jpn_keyword, eng_keyword):
        if os.path.exists(os.path.join(self.new_dir_path, jpn_keyword)) == False:
            os.mkdir(os.path.join(self.new_dir_path, jpn_keyword))
        if os.path.isdir(os.path.join(self.new_dir_path, eng_keyword)) == False:
            os.mkdir(os.path.join(self.new_dir_path, eng_keyword))
        """if os.path.isdir(os.path.join(self.new_dir_path, ind_keyword)) == False:
            os.mkdir(os.path.join(self.new_dir_path, ind_keyword))"""

    # 確信度を計算
    def confidence_cal(self, c, str_label_list, feature_vector):
        confi_list_a = [] # "[1. 0.]"ラベルの特徴ベクトルを格納
        confi_list_b = [] # "[0. 1.]"ラベルの特徴ベクトルを格納
        result_list_a = [] # "[1. 0.]"ラベルの総当たりでの類似度を格納
        result_list_b = [] # "[0. 1.]"ラベルの総当たりでの類似度を格納
        confidence_feature_dict = {i:[] for i in c.keys()}
        # feature_vector(特徴ベクトル)を辞書にラベルと対応づけて格納
        for i in range(len(feature_vector)):
            #confidence_feature_dict[str_label_list[i]].append(np.reshape(feature_vector[i], 7*7*512))
            confidence_feature_dict[str_label_list[i]].append(np.reshape(feature_vector[i], 4*4*512))

        # ラベルごとに特徴ベクトルをリストに格納
        for a in confidence_feature_dict["[1. 0.]"]:
            confi_list_a.append(a)
        # リストの中身を総当たりで類似度を計算し、リストに格納
        for confi_match_a in itertools.combinations(confi_list_a, 2):
            result_list_a.append(self.cos_sim(confi_match_a[0], confi_match_a[1]))

        number = 1
        for result in result_list_a:
            number += 1
        print('確信度(平均):', self.average(result_list_a))

        # ラベルごとに特徴ベクトルをリストに格納
        for b in confidence_feature_dict["[0. 1.]"]:
            confi_list_b.append(b)
        # リストの中身を総当たりで類似度を計算し、リストに格納
        for confi_match_b in itertools.combinations(confi_list_b, 2):
            result_list_b.append(self.cos_sim(confi_match_b[0], confi_match_b[1]))
        print('確信度(平均):', self.average(result_list_b))

        book = openpyxl.load_workbook("confidence_test.xlsx")
        sheet = book.worksheets[0]
        max_row = sheet.max_row # シートの最後の行を取得

        sheet.cell(row = max_row + 1, column=1).value = max_row
        sheet.cell(row = max_row + 1, column=2).value = self.average(result_list_a)
        sheet.cell(row = max_row + 1, column=3).value = self.average(result_list_b)

        # 保存
        book.save("confidence_test.xlsx")
        # 終了
        book.close()

    #ファイルの名前取得
    def filename_get(self, jpn_keyword, eng_keyword):
        filenames_japanese = os.listdir(os.path.join(self.dir_path, jpn_keyword, ""))
        filenames_english = os.listdir(os.path.join(self.dir_path, eng_keyword, ""))
        #filenames_indlish = os.listdir(os.path.join(self.dir_path, ind_keyword, ""))

        for filename_jpn in filenames_japanese:
            self.imread_gray(filename_jpn, os.path.join(self.dir_path, jpn_keyword, ""), os.path.join(self.new_dir_path, jpn_keyword, ""))
        for filename_eng in filenames_english:
            self.imread_gray(filename_eng, os.path.join(self.dir_path, eng_keyword, ""), os.path.join(self.new_dir_path, eng_keyword, ""))
        """for filename_ind in filenames_indlish:
            self.imread_gray(filename_ind, os.path.join(self.dir_path, ind_keyword, ""), os.path.join(self.new_dir_path, ind_keyword, ""))"""

    def outlier_image(self, vector_dict, labels):
        for vector, label in zip(vector_dict[0], labels):
            print('{}, ベクトル{}'.format(label, vector))

    #外れ値検出
    def isolation(self, vector_dict):
        data = self.vector_list_get(vector_dict)
        clf = IsolationForest(n_estimators=100,
                            contamination='auto',
                            #behaviour='new',
                            random_state=42)
        # 学習用データを学習させる
        clf.fit(data)
        # 検証用データを分類する
        pred = clf.predict(data)
        
        self.outlier_image(vector_dict, self.weight_list_get(pred))
        return self.weighted_average(data, self.weight_list_get(pred))

    #DBSCAN
    def dbscan(self, vector_dict):
        data = self.vector_list_get(vector_dict)
        db = DBSCAN(eps=1000, min_samples=2).fit(data)
        labels = db.labels_
        
        self.outlier_image(vector_dict, self.weight_list_get(labels))

        return self.weighted_average(data, self.weight_list_get(labels))

    #マハラノビス距離を計算
    def mahalanobis(self, data):
        # データ集合の平均値mean, データ集合の共分散行列(分散を集めたもの)cov
        for i in range(len(data)):
            mean += data[i]
        mean = mean / len(data)
        data_m = data - mean
        cov = np.cov(np.array(data).T)
        result = []
        for i in range(len(data)):
            result.append(distance.mahalanobis(data[i], mean, np.linalg.pinv(cov)))

        return result

    # パスの結合
    def path_join(self, path, dir):
        return os.path.join(path, dir, "")

    # フォルダの中身を全て取得
    def dir_get(self, keyword, path):
        return os.listdir(self.path_join(path, keyword))

    #加重平均
    def weighted_average(self, vector_list, weight_list):
        weighted_vector = np.zeros(7*7*512)
        for vector, weight in zip(vector_list, weight_list):
            weighted_vector += vector*weight

        return weighted_vector / sum(weight_list)

    def vector_list_get(self, vector_dict):
        vector_list = []

        for vector_lists in vector_dict.values():
            for vector in vector_lists:
                vector_list.append(vector[1])

        return vector_list

    def weight_list_get(self, label_list):
        c = collections.Counter(label_list)
        weight_list = []
        for label in label_list:
            weight_list.append(c[label])

        return weight_list

    def vgg_vector(self, keyword, include_top_label):
        # VGG16のモデル
        model = VGG16(include_top=include_top_label, weights='imagenet') # 全結合層なし
        vector_dict = defaultdict(list) # ベクトルを入れる辞書(key='数字'クラス数, value=(画像名, 'ベクトル'7*7*512))

        # キーワードのフォルダ内にある画像フォルダを取得
        folders = self.dir_get(keyword, self.new_dir_path + '/')

        # 辞書のkeyにする数字, クラスの数
        key_index = 0

        # フォルダのリストをスライスして, フォルダ内の画像一覧を取得
        for folder in folders:
            images = self.dir_get(folder, self.new_dir_path + '/' + keyword + '/') # フォルダ内の画像の一覧を取得
            # 画像の一覧をスライスして, 一つずつベクトルを取得
            for image in images:
                # 画像を取り込む
                try:
                    img = load_img(self.new_dir_path + '/' + keyword + '/' + folder + '/' + image, target_size=(224, 224))
                    # 画像をベクトルに変換
                    ary = img_to_array(img)
                    # VGG16で特徴ベクトルに変換 'shape = (7, 7, 512)'
                    vector = model.predict(preprocess_input(np.expand_dims(ary, axis=0)))
                    # 辞書にベクトルを格納 key=数字, value=(画像名, ベクトル)
                    vector_dict[key_index].append([keyword + '/' + folder + '/' + image, np.reshape(vector, 7*7*512)])
                except:
                    print(self.new_dir_path + '/' + keyword + '/' + folder + '/' + image)

                # ベクトルを保存
                #self.save_vector(folder + '/' + image, np.reshape(vector, 7*7*512))

            # クラスが変わるごとにkeyの値を増やす
            key_index += 1

        return vector_dict

    #文化差検出
    def cd_detect(self, jpn_keyword, eng_keyword, detect_model):
        #self.change_gray(jpn_keyword, self.dir_path)
        #self.change_gray(eng_keyword, self.dir_path)

        if detect_model == 'predict':
            jpn_vector_dict = self.vgg_vector(jpn_keyword, True)
            eng_vector_dict = self.vgg_vector(eng_keyword, True)
            
            print(jpn_vector_dict[0])

        else:
            jpn_vector_dict = self.vgg_vector(jpn_keyword, False)
            eng_vector_dict = self.vgg_vector(eng_keyword, False)
            print(jpn_vector_dict[0])

            if detect_model == "Avg":
                jpn_vector = self.simple_avg(jpn_vector_dict)
                eng_vector = self.simple_avg(eng_vector_dict)

            elif detect_model == 'Kmeans':
                jpn_vector, eng_vector = self.kmeans(jpn_vector_dict, eng_vector_dict)

            elif detect_model == 'DBSCAN':
                jpn_vector = self.dbscan(jpn_vector_dict)
                eng_vector = self.dbscan(eng_vector_dict)

            elif detect_model == 'Isolation':
                jpn_vector = self.isolation(jpn_vector_dict)
                eng_vector = self.isolation(eng_vector_dict)

            elif detect_model == '重心':
                jpn_vector = self.median_vector(jpn_vector_dict, 'Jpn')
                eng_vector = self.median_vector(eng_vector_dict, 'Eng')

        #print(self.cos_sim(jpn_vector, eng_vector))
        #self.write_result(self.cos_sim(jpn_vector, eng_vector), jpn_keyword, eng_keyword)

        return

    def simple_avg(self, vector_dict):
        count = 0
        vector_avg = np.zeros(7*7*512)
        for vector_lists in vector_dict.values():
            for vector in vector_lists:
                vector_avg += vector[1]
                count += 1
        vector_avg /= count

        return vector_avg

    def kmeans(self, jpn_vector_dict, eng_vector_dict):
        clf = KMeans(n_clusters = 2)
        kmeans_vector = np.zeros(7*7*512)

        jpn_vector_list = self.vector_list_get(jpn_vector_dict)
        eng_vector_list = self.vector_list_get(eng_vector_dict)

        kmeans_jpn_label = clf.fit_predict(jpn_vector_list)
        kmeans_eng_label = clf.fit_predict(eng_vector_list)
        
        self.outlier_image(jpn_vector_dict, self.weight_list_get(kmeans_jpn_label))
        self.outlier_image(eng_vector_dict, self.weight_list_get(kmeans_eng_label))

        return self.weighted_average(jpn_vector_list, self.weight_list_get(kmeans_jpn_label)), self.weighted_average(eng_vector_list, self.weight_list_get(kmeans_eng_label))

    # 重心の重心を求める
    def median_vector(self, vector_dict, flag):
        median_dict = {i:np.zeros(7*7*512) for i in vector_dict.keys()} # 重心を格納する辞書
        median_median = np.zeros(7*7*512) # 重心の重心を格納

        for key in vector_dict.keys():
            for vector in vector_dict[key]:
                median_dict[key] += vector[1]
            median_dict[key] /= len(vector_dict[key])

        for median in median_dict.values():
            median_median += median
        median_median /= len(median_dict.values())

        print('重心の重心{}'.format(median_median))

        return median_median

    """#文化差検出
    def cd_detect(self, jpn_keyword, eng_keyword, detect_model):
        print(detect_model)
        #self.filename_get(jpn_keyword, eng_keyword)
        batch_size = 32 #バッチサイズ
        if detect_model == 'predict':
            #VGG16のモデル
            model_confidence = VGG16(include_top=True, weights='imagenet', input_shape=(224,224,3))
            image_data_generator = ImageDataGenerator()
            train_data = image_data_generator.flow_from_directory(
                'downloads/grayscale',
                target_size = (224, 224), #予測するならこっち
                batch_size = batch_size,
                class_mode = 'categorical',
                classes = {jpn_keyword, eng_keyword},
                shuffle = False
            )

            counter = 0
            label_list = [] # ラベルを格納するリスト ラベルは[1. 0.] と [0. 1.]
            for input, label in train_data:
                counter += 1
                label_list.append(label)
                if batch_size * counter >= 20:
                    break

            str_label_list = [str(i) for i in label_list[0]] # ラベルをstr型に変換
            c = collections.Counter(str_label_list) # ラベルごとに個数を数える
            feature_vector_average_dict = {i:np.zeros(7*7*512) for i in c.keys()} # 各ラベルをキーにして辞書型を作る 平均特徴ベクトルをいれる辞書型 初期値は0
            label_count_dict = {i:np.zeros(1) for i in c.keys()} # 各ラベルをキーにして辞書型を作る 平均特徴ベクトルをいれる辞書型 初期値は0
            feature_vector = model.predict_generator(train_data, 1, verbose=1)

            #予測による重みを決める
            vectors_data_list = []
            predict_vectors = model_confidence.predict_generator(train_data, 1, verbose=1)
            predict_result = []
            for pred, label, vector in zip(predict_vectors, str_label_list, feature_vector):
                vector = np.reshape(vector, 7*7*512)
                pred = np.expand_dims(pred, axis=0)
                results = decode_predictions(pred, top=1)[0]
                predict_result.append([results[0][1], label])
                vectors_data_list.append([label, vector, results[0][1]])
            
            predict_list1 = []
            predict_list2 = []
            for predict in predict_result:
                if predict[1] == str_label_list[0]:
                    predict_list1.append(predict[0])
                else:
                    predict_list2.append(predict[0])

            self.confidence_cal(c, str_label_list, feature_vector)
            for i in range(len(vectors_data_list)):
                if vectors_data_list[i][0] == str_label_list[0]:
                    feature_vector_average_dict[str_label_list[i]] += vectors_data_list[i][1] * predict_list1.count(vectors_data_list[i][2])
                    label_count_dict[str_label_list[i]] += predict_list1.count(vectors_data_list[i][2])
                else:
                    feature_vector_average_dict[str_label_list[i]] += vectors_data_list[i][1] * predict_list2.count(vectors_data_list[i][2])
                    label_count_dict[str_label_list[i]] += predict_list2.count(vectors_data_list[i][2])
            for key in c.keys():
                feature_vector_average_dict[key] /= label_count_dict[key]

            cos_similarity = self.cos_sim(feature_vector_average_dict["[1. 0.]"], feature_vector_average_dict["[0. 1.]"])
            self.print_cossim(cos_similarity)

            #self.write_result(cos_similarity, jpn_keyword, eng_keyword)

        else:
            #VGG16のモデル
            model = VGG16(include_top=False, weights='imagenet') #全結合層なし
            image_data_generator = ImageDataGenerator()
            train_data = image_data_generator.flow_from_directory(
                'downloads/grayscale',
                target_size = (150, 150), #予測しない
                batch_size = batch_size,
                class_mode = 'categorical',
                classes = {jpn_keyword, eng_keyword},
                shuffle = False
                )

            counter = 0
            label_list = [] # ラベルを格納するリスト ラベルは[1. 0.] と [0. 1.]
            for input, label in train_data:
                counter += 1
                label_list.append(label)
                if batch_size * counter >= 20:
                    break

            str_label_list = [str(i) for i in label_list[0]] # ラベルをstr型に変換
            c = collections.Counter(str_label_list) # ラベルごとに個数を数える
            feature_vector = model.predict_generator(train_data, 1, verbose=1)
            self.confidence_cal(c, str_label_list, feature_vector)

            if detect_model == 'average':
                feature_vector_average_dict = self.cd_detect_average(c, str_label_list, feature_vector)
            elif detect_model == 'Kmeans':
                feature_vector_average_dict = self.cd_detect_kmeans(c, str_label_list, feature_vector)
            elif detect_model == 'Kmeans-outlier':
                feature_vector_average_dict = self.cd_detect_kmeans_outlier(c, str_label_list, feature_vector)
            else:
                feature_vector_average_dict = {i:np.zeros(4*4*512) for i in c.keys()} # 各ラベルをキーにして辞書型を作る 平均特徴ベクトルをいれる辞書型 初期値は0
                label_count_dict = {i:np.zeros(1) for i in c.keys()} # 各ラベルをキーにして辞書型を作る 平均特徴ベクトルをいれる辞書型 初期値は0

                # 外れ値検出
                outlier_list = [] #Isolationでクラスタリングするためのベクトルを入れる多次元リスト[ラベル, ベクトル]
                outlier_class1 = [] #Kmeans法でクラスタリングするためのベクトルを入れるリスト
                outlier_class2 = [] #Kmeans法でクラスタリングするためのベクトルを入れるリスト
                pred_list = []
                for vector, label in zip(feature_vector, str_label_list):
                    vector = np.reshape(vector, 4*4*512)
                    outlier_list.append([label,vector])
                    if label == str_label_list[0]:
                        outlier_class1.append(vector)
                    else:
                        outlier_class2.append(vector)
                if detect_model == 'Isolation':
                    pred_class1 = self.isolation(outlier_class1)
                    pred_class2 = self.isolation(outlier_class2)
                elif detect_model == 'DBSCAN':
                    pred_class1 = self.dbscan(outlier_class1)
                    pred_class2 = self.dbscan(outlier_class2)
                elif detect_model == 'mahalanobis':
                    pred_class1 = self.mahalanobis(outlier_class1)
                    pred_class2 = self.mahalanobis(outlier_class2)

                pred_list.extend(self.weight_cal(pred_class1))
                pred_list.extend(self.weight_cal(pred_class2))

                feature_vector_average_dict = self.weighted_average(outlier_list, pred_list, str_label_list, c)

            cos_similarity = self.cos_sim(feature_vector_average_dict["[1. 0.]"], feature_vector_average_dict["[0. 1.]"])
            self.print_cossim(cos_similarity)
            #self.write_result(cos_similarity, jpn_keyword, eng_keyword)

    # 単純平均で平均特徴ベクトルを生成
    def cd_detect_average(self, c, str_label_list, feature_vector):
        train_feature_dict = {i:[] for i in c.keys()} # 各ラベルをキーにして辞書型を作る keyにラベル, valueにそれぞれの特徴量をいれる
        feature_vector_average_dict = {i:np.zeros(4*4*512) for i in c.keys()} # 各ラベルをキーにして辞書型を作る 平均特徴ベクトルをいれる辞書型 初期値は0

        for i in range(len(feature_vector)):
            train_feature_dict[str_label_list[i]].append(np.reshape(feature_vector[i], 4*4*512))
            feature_vector_average_dict[str_label_list[i]] += train_feature_dict[str_label_list[i]][-1]

        for key in c.keys():
            feature_vector_average_dict[key] /= c[key]

        return feature_vector_average_dict

    #Kmeansを使って平均特徴ベクトルを生成
    def cd_detect_kmeans(self, c, str_label_list, feature_vector):
        clf = KMeans(n_clusters = 2)
        feature_vector_average_dict = {i:np.zeros(4*4*512) for i in c.keys()} # 各ラベルをキーにして辞書型を作る 平均特徴ベクトルをいれる辞書型 初期値は0

        # Kmeans法によるクラスタリング
        kmeans_vector_list = [] #Kmeans法でクラスタリングするためのベクトルを入れる多次元リスト[ラベル, ベクトル]
        kmeans_vector_class1 = [] #Kmeans法でクラスタリングするためのベクトルを入れるリスト
        kmeans_vector_class2 = [] #Kmeans法でクラスタリングするためのベクトルを入れるリスト
        pred_list = []
        for vector, label in zip(feature_vector, str_label_list):
            vector = np.reshape(vector, 4*4*512)
            kmeans_vector_list.append([label,vector])
            if label == str_label_list[0]:
                kmeans_vector_class1.append(vector)
            else:
                kmeans_vector_class2.append(vector)
        pred_class1 = clf.fit_predict(kmeans_vector_class1)
        pred_class2 = clf.fit_predict(kmeans_vector_class2)

        pred_list.extend(self.weight_cal(pred_class1))
        pred_list.extend(self.weight_cal(pred_class2))
        
        print(pred_list)
        feature_vector_average_dict = self.weighted_average(kmeans_vector_list, pred_list, str_label_list, c)

        return feature_vector_average_dict

    #Kmeansを使って平均特徴ベクトルを生成
    def cd_detect_kmeans_outlier(self, c, str_label_list, feature_vector):
        clf = KMeans(n_clusters = 2)
        feature_vector_average_dict = {i:np.zeros(4*4*512) for i in c.keys()} # 各ラベルをキーにして辞書型を作る 平均特徴ベクトルをいれる辞書型 初期値は0

        # Kmeans法によるクラスタリング
        kmeans_vector_list = [] #Kmeans法でクラスタリングするためのベクトルを入れる多次元リスト[ラベル, ベクトル]
        kmeans_vector_class1 = [] #Kmeans法でクラスタリングするためのベクトルを入れるリスト
        kmeans_vector_class2 = [] #Kmeans法でクラスタリングするためのベクトルを入れるリスト
        pred_list = []
        for vector, label in zip(feature_vector, str_label_list):
            vector = np.reshape(vector, 4*4*512)
            kmeans_vector_list.append([label,vector])
            if label == str_label_list[0]:
                kmeans_vector_class1.append(vector)
            else:
                kmeans_vector_class2.append(vector)

        outlier1 = self.isolation(kmeans_vector_class1)
        outlier2 = self.isolation(kmeans_vector_class2)
        if -1 in outlier1:
            pred_class1 = clf.fit_predict(kmeans_vector_class1)
        else:
            pred_class1 = np.ones(len(kmeans_vector_class1), dtype='int32')
        if -1 in outlier2:
            pred_class2 = clf.fit_predict(kmeans_vector_class2)
        else:
            pred_class2 = np.ones(len(kmeans_vector_class2), dtype='int32')

        pred_list.extend(self.weight_cal(pred_class1))
        pred_list.extend(self.weight_cal(pred_class2))

        feature_vector_average_dict = self.weighted_average(kmeans_vector_list, pred_list, str_label_list, c)

        return feature_vector_average_dict"""

detect_model = ['Avg', 'Kmeans', 'predict', 'Isolation', 'mahalanobis', 'DBSCAN', 'Kmeans-outlier']
cdd = Cultural_Difference_Detector()
cdd.get_synset()

for key in range(231, 232): # 1000個一気にやるとメモリ不足になるから100ずつ区切ってやる
    print('========================================')
    print('ID:{},キーワード:{}'.format(key, cdd.jpn_keyword_dict[key]))
    print('========================================')
    #cdd.image_confirm(key)
    #cdd.folder_confirm(os.path.join(cdd.new_dir_path, cdd.jpn_keyword_dict[key]))
    #cdd.folder_confirm(os.path.join(cdd.new_dir_path, cdd.eng_keyword_dict[key]))
    cdd.cd_detect(cdd.jpn_keyword_dict[key], cdd.eng_keyword_dict[key], detect_model[2])
