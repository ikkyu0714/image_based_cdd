"""
https://aidiary.hatenablog.com/entry/20170110/1484057655 参考
"""
#from re import T
from copyreg import pickle
from tensorflow.keras.applications import VGG16
from tensorflow.keras.applications.vgg16 import VGG16, decode_predictions, preprocess_input
import os, zipfile, requests, argparse, json, urllib, urllib.request, collections, pickle, gc
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from tensorflow.keras.preprocessing.image import load_img, img_to_array
from PIL import Image
from tensorflow.python.eager.tape import delete_trace
#from google_images_download_patch.google_images_download.google_images_download import googleimagesdownload
from google_images_download.google_images_download import googleimagesdownload
import openpyxl
import itertools
import pprint
from PIL import ImageFile
from collections import defaultdict
from sklearn.manifold import TSNE
ImageFile.LOAD_TRUNCATED_IMAGES = True

# 文化差検出クラス
class Cultural_Defference_Detector():
    def __init__(self):
        self.readfile = 'synset_改良.xlsx' # 読み込むsynsetのリストを記録したファイル
        self.filename='../../研究/類似度結果/重心の重心結果1103.xlsx' # 結果を書き込むExcelのパス
        self.jpn_keyword_dict = {} # キーワードのリストを取得する辞書（日本語）
        self.eng_keyword_dict = {} # キーワードのリストを取得する辞書（英語）
        self.jpn_hyper_dict = {}
        self.eng_hyper_dict = {}
        self.ind_keyword_dict = {} #
        self.new_dir_path = "downloads" # グレースケール化した後のフォルダ
        self.dir_path = "downloads" # 画像が保存されているフォルダ

    # Synsetを取得する
    def get_synset(self):
        key = 1
        book = openpyxl.load_workbook(self.readfile)
        sheet = book.worksheets[0]

        for row in sheet.iter_rows(min_row=2, min_col=2):
            value = []
            for cell in row:
                values = cell.value
                #values = values.replace(',', '') # 検証の時に使う
                value.append(values.replace('_', ' '))

            self.jpn_hyper_dict[key] = value[0]
            self.jpn_keyword_dict[key] = value[1]
            self.eng_hyper_dict[key] = value[2]
            self.eng_keyword_dict[key] = value[3]
            key += 1

        # 上書き保存（読み込んだのと同じ名前を指定）
        book.save(self.readfile)

        book.close()

    # 結果をエクセルに書き込む
    def write_result(self, result, jpn_keyword, eng_keyword):
        book = openpyxl.load_workbook(self.filename)
        sheet = book.worksheets[0]
        max_row = sheet.max_row # シートの最後の行を取得

        # 最後の一つ下の行に (3列目:日本語のキーワード, 4列目:英語のキーワード, 5列目:cos類似度) を書き込む
        sheet.cell(row = max_row + 1, column=3).value = jpn_keyword
        sheet.cell(row = max_row + 1, column=4).value = eng_keyword
        sheet.cell(row = max_row + 1, column=5).value = result
        # 保存
        book.save(self.filename)
        # 終了
        book.close()

    # cos類似度を計算する
    def cos_sim(self, v1, v2):
        print('分子{}'.format(np.dot(v1, v2)))
        print('分母{},{}'.format(np.linalg.norm(v1), np.linalg.norm(v2)))
        return np.dot(v1, v2) / (np.linalg.norm(v1) * np.linalg.norm(v2))

    # 画像をグレースケール化
    def imread_gray(self, image_name, path1, path2):
        gamma22LUT  = [pow(x/255.0, 2.2)*255 for x in range(256)] * 3
        gamma045LUT = [pow(x/255.0, 1.0/2.2)*255 for x in range(256)]
        img = Image.open(path1 + image_name)
        img_rgb = img.convert("RGB")
        img_rgbL = img_rgb.point(gamma22LUT)
        img_grayL = img_rgbL.convert("L")
        img_gray = img_grayL.point(gamma045LUT)
        img_gray.save(path2 + image_name)

    # 画像検索
    def image_search(self, keywords, hypernyms, exist_num, key):
        keyword = keywords.split(", ")
        hypernym = hypernyms.split(", ")
        for word in keyword:
            for hyper in hypernym:
                target = word + " " + hyper
                lang_flag = target.isascii()
                if lang_flag == False:
                    language = "Japanese"
                else:
                    language = "English"
                print('{}:{}'.format(key, target))
                limit = 10 - exist_num
                try:
                    response = googleimagesdownload()  #responseオブジェクトの生成
                    arguments = {"keywords":target,  # 検索キーワード
                                "limit":limit,  # ダウンロードする画像の数(デフォルト100)
                                "format":"jpg",
                                "language":language,
                                "output_directory":"downloads/" + keywords,
                                "chromedriver":"chromedriver.exe"
                                }
                    response.download(arguments)   # argumentsをresponseオブジェクトのdownloadメソッドに渡す
                except Exception as e:
                    print("Caught error:")
                    print(e)

    # 画像があるか確認
    def image_confirm(self, key):
        if os.path.isdir(os.path.join(self.dir_path, self.eng_keyword_dict[key])) == False:
            self.image_search(self.eng_keyword_dict[key], self.eng_hyper_dict[key], 0, key)
        else:
            folders = os.listdir(os.path.join(self.dir_path, self.eng_keyword_dict[key]))
            for folder in folders:
                if self.file_or_folder(folder) == False:
                    path = os.path.join(self.dir_path, self.eng_keyword_dict[key])
                    files = os.listdir(os.path.join(path, folder))
                    files_file = [f for f in files if os.path.isfile(os.path.join(os.path.join(path, folder), f))]
                    if len(files_file) < 10:
                        self.image_search(self.eng_keyword_dict[key], self.eng_hyper_dict[key], len(files_file), key)

        if os.path.isdir(os.path.join(self.dir_path, self.jpn_keyword_dict[key])) == False:
            self.image_search(self.jpn_keyword_dict[key], self.jpn_hyper_dict[key], 0, key)
        else:
            folders = os.listdir(os.path.join(self.dir_path, self.jpn_keyword_dict[key]))
            for folder in folders:
                if self.file_or_folder(folder) == False:
                    path = os.path.join(self.dir_path, self.jpn_keyword_dict[key])
                    files = os.listdir(os.path.join(path, folder))
                    files_file = [f for f in files if os.path.isfile(os.path.join(os.path.join(path, folder), f))]
                    if len(files_file) < 10:
                        self.image_search(self.jpn_keyword_dict[key], self.jpn_hyper_dict[key], len(files_file), key)

    # フォルダーがあるか確認
    def folder_confirm(self, path):
        if os.path.exists(path) == False:
            os.mkdir(path)

    # パスの結合
    def path_join(self, path, dir):
        return os.path.join(path, dir, "")

    def change_gray(self, keyword, path):
        # フォルダの中身を取得
        dirnames = self.dir_get(keyword, path)
        path = self.path_join(path, keyword)

        for dirname in dirnames:
            new_path = keyword + '/' + dirname
            # 中身がファイルならグレースケール化
            if self.file_or_folder(os.path.join(path, dirname)):
                self.folder_confirm(os.path.join(self.new_dir_path, new_path))
                self.imread_gray(dirname, path, self.path_join(self.new_dir_path, new_path))
            # フォルダなら再帰呼び出し
            else:
                # フォルダの中身を取得
                filenames = self.dir_get(dirname, path)
                path2 = self.path_join(path, dirname)
                for filename in filenames:
                    new_path2 = new_path + '/' + filename
                    # 中身がファイルならグレースケール化
                    if self.file_or_folder(os.path.join(path2, dirname)):
                        self.folder_confirm(os.path.join(self.new_dir_path, new_path2))
                        self.imread_gray(dirname, path2, self.path_join(self.new_dir_path, new_path2))

    # ファイルかフォルダかを判断
    def file_or_folder(self, path):
        if os.path.isfile(path):
            return True
        else:
            return False

    # フォルダの中身を全て取得
    def dir_get(self, keyword, path):
        return os.listdir(self.path_join(path, keyword))

    def vgg_vector(self, keyword):
        # VGG16のモデル
        model = VGG16(include_top=False, weights='imagenet') # 全結合層なし
        vector_dict = defaultdict(list) # ベクトルを入れる辞書(key='数字'クラス数, value=(画像名, 'ベクトル'7*7*512))

        # キーワードのフォルダ内にある画像フォルダを取得
        folders = self.dir_get(keyword, self.new_dir_path + '/')

        # 辞書のkeyにする数字, クラスの数
        key_index = 0

        # フォルダのリストをスライスして, フォルダ内の画像一覧を取得
        for folder in folders:
            images = self.dir_get(folder, self.new_dir_path + '/' + keyword + '/') # フォルダ内の画像の一覧を取得
            # 画像の一覧をスライスして, 一つずつベクトルを取得
            for image in images:
                # 画像を取り込む
                try:
                    img = load_img(self.new_dir_path + '/' + keyword + '/' + folder + '/' + image, target_size=(224, 224))
                    # 画像をベクトルに変換
                    ary = img_to_array(img)
                    # VGG16で特徴ベクトルに変換 'shape = (7, 7, 512)'
                    vector = model.predict(preprocess_input(np.expand_dims(ary, axis=0)))
                    # 辞書にベクトルを格納 key=数字, value=(画像名, ベクトル)
                    vector_dict[key_index].append([keyword + '/' + folder + '/' + image, np.reshape(vector, 7*7*512)])
                except:
                    print(self.new_dir_path + '/' + keyword + '/' + folder + '/' + image)

                # ベクトルを保存
                #self.save_vector(folder + '/' + image, np.reshape(vector, 7*7*512))

            # クラスが変わるごとにkeyの値を増やす
            key_index += 1

        return vector_dict

    #文化差検出
    def cd_detect(self, jpn_keyword, eng_keyword):
        #self.change_gray(jpn_keyword, self.dir_path)
        #self.change_gray(eng_keyword, self.dir_path)

        jpn_vector_dict = self.vgg_vector(jpn_keyword)
        eng_vector_dict = self.vgg_vector(eng_keyword)

        jpn_vector = self.median_vector(jpn_vector_dict, 'Jpn')
        eng_vector = self.median_vector(eng_vector_dict, 'Eng')

        print(self.cos_sim(jpn_vector, eng_vector))
        #self.write_result(self.cos_sim(jpn_vector, eng_vector), jpn_keyword, eng_keyword)

        return

    def distance_cal(self, vector1, vector2):
        dist = np.sqrt(np.sum(np.square(vector1-vector2)))
        return dist
    
    def save_vector(self, imagename, vector):
        f = open('vector.dat', 'at')
        f.write(imagename)
        f.write('\n')
        f.write(np.linalg.norm(vector,ord=2))
        f.write('\n')
        f.write(vector)
        f.close()

    # 重心の重心を求める
    def median_vector(self, vector_dict, flag):
        median_dict = {i:np.zeros(7*7*512) for i in vector_dict.keys()} # 重心を格納する辞書
        median_median = np.zeros(7*7*512) # 重心の重心を格納

        for key in vector_dict.keys():
            for vector in vector_dict[key]:
                median_dict[key] += vector[1]
            median_dict[key] /= len(vector_dict[key])

        for median in median_dict.values():
            median_median += median
        median_median /= len(median_dict.values())

        """for index in np.where(median_median==0.0):
            print(index)"""
        #print(value)
        print('重心の重心{}'.format(median_median))


        return median_median

cdd = Cultural_Defference_Detector()
cdd.get_synset()

for key in range(1,2): # 1000個一気にやるとメモリ不足になるから100ずつ区切ってやる
    print('========================================')
    print('ID:{},キーワード:{}'.format(key,cdd.jpn_keyword_dict[key]))
    print('========================================')
    cdd.image_confirm(key)
    #cdd.folder_confirm(os.path.join(cdd.new_dir_path, cdd.jpn_keyword_dict[key]))
    #cdd.folder_confirm(os.path.join(cdd.new_dir_path, cdd.eng_keyword_dict[key]))
    cdd.cd_detect(cdd.jpn_keyword_dict[key], cdd.eng_keyword_dict[key])
    if key % 10 == 0:
        print(key)
        gc.collect()
